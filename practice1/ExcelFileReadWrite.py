import openpyxl

workbook = openpyxl.load_workbook("Test.xlsx")

sheet = workbook["FirstSheet"]

#sheet = workbook.active # This will return actice Sheet

print("Max Row = ", sheet.max_row)      # This will print Max Row Number
print("Max Column = ", sheet.max_column)  # This will print max Column

for rowno in range(1, sheet.max_row+1):
    for colno in range(1, sheet.max_column+1):
        print(sheet.cell(row=rowno, column=colno).value, end=" ")
    print("\n")

mydict = {}

rowno = 3

for colno in range(1, sheet.max_column+1):
    mydict[sheet.cell(row=1, column=colno).value] = sheet.cell(row=rowno, column=colno).value

print(mydict)

sheet.cell(row=2, column=1).value = "NageshBappa"   # Writting data into File

print(sheet.cell(row=2, column=1).value)

sheet.cell(row=2, column=1).value = "Nagesh"

workbook.close()

